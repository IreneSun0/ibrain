#!/usr/bin/env python3
"""ingest_xlsx.py — deterministic importer for a 4-sheet learning-map workbook.

Behaviour (see xlsx-import-report.md after each run):
- 术语总表   → one concept page per row (entity-like rows → 03_ENTITIES/protocols-networks/)
- 学习地图   → 10_LEARNING/curriculum/curriculum-source-map.md (verbatim, source of truth for curriculum)
- 生态游戏版图 → 06_RELATIONSHIPS/ecosystem-maps/ecosystem-roles-map.md (verbatim)
- 机构对话速查 → 10_LEARNING/expert-questions/institutional-conversation-cheatsheet.md (verbatim)
- unique Source URLs → source notes in 07_RESEARCH/sources/

Idempotency / merge rules:
- A generated page carries `import_origin: xlsx-learning-map` in frontmatter.
- Re-run: page exists with that marker → rewritten (importer owns it).
- Page exists WITHOUT the marker (human/agent enriched or pre-existing richer note)
  → NEVER overwritten; row is listed in the report under "requires human merge".
- After manual enrichment, set `import_origin: xlsx-learning-map+manual` to lock the page.

No LLM anywhere in this file. IDs/slugs/hashes are code.
"""
from __future__ import annotations

import hashlib
import os
import sys
from datetime import date
from pathlib import Path
from urllib.parse import urlparse

sys.path.insert(0, str(Path(__file__).resolve().parent))
from brainlib import OPS_ROOT, build_frontmatter, load_note, sha256_file, slugify, vault_root

TODAY = "2026-08-26"
SEED_XLSX = Path(os.environ.get(
    "IBRAIN_XLSX_PATH",
    os.environ.get("IBRAIN_WORKBOOK", str(OPS_ROOT / "_seed" / "learning-map.xlsx")),
))
WORKBOOK_SOURCE_ID = "source:2026-08-26-industry-learning-map-xlsx"
MARKER = "xlsx-learning-map"

# ── deterministic slug overrides (English Term → canonical slug) ──────────────
SLUG_OVERRIDES = {
    "CLOB": "central-limit-order-book",
    "Perpetual Futures / Perp": "perpetual-futures",
    "Binary/Digital Option": "binary-option",
    "JUST / JustLend DAO": "justlend",
    "Bandwidth (TRON)": "tron-bandwidth",
    "Energy (TRON)": "tron-energy",
    "TRON Power": "tron-power",
    "Energy Delegation": "tron-energy-delegation",
    "ERC Standard": "erc-standards",
    "VaR": "value-at-risk",
    "AMM": "automated-market-maker",
    "RFQ": "request-for-quote",
    "Smart Order Routing": "smart-order-routing",
    "KYT": "know-your-transaction",
    "KYC": "know-your-customer",
    "AML": "anti-money-laundering",
    "PoW": "proof-of-work",
    "PoS": "proof-of-stake",
    "DPoS": "delegated-proof-of-stake",
    "L1": "layer-1",
    "L2": "layer-2",
    "EVM": "ethereum-virtual-machine",
    "CEX": "centralized-exchange",
    "DEX": "decentralized-exchange",
    "OTC": "over-the-counter",
    "BTTC": "bittorrent-chain",
    "Forward": "forward-contract",
    "Futures": "futures-contract",
    "Option": "option",
    "Swap": "swap",
    "Event VaR": "event-var",
    "Fully Collateralized": "fully-collateralized-market",
    "On-chain": "on-chain",
    "Off-chain": "off-chain",
    "Hybrid Exchange": "hybrid-exchange-architecture",
    "TRC-20": "trc-20",
    "ERC-20": "erc-20",
    "ERC-721": "erc-721",
    "ERC-1155": "erc-1155",
}

# Terms that are actually named protocols/products, not abstract concepts.
ENTITY_TERMS = {
    "SunPump": ("protocol", "03_ENTITIES/protocols-networks", "protocol-network", "sunpump"),
    "JUST / JustLend DAO": ("protocol", "03_ENTITIES/protocols-networks", "protocol-network", "justlend"),
    "BTTC": ("protocol", "03_ENTITIES/protocols-networks", "protocol-network", "bittorrent-chain"),
    "WINkLink": ("protocol", "03_ENTITIES/protocols-networks", "protocol-network", "winklink"),
}

# 分类 → (concept subfolder, domains list)
CATEGORY_MAP = {
    "市场结构": ("financial-markets", ["financial-markets"]),
    "市场微观结构": ("market-microstructure", ["market-microstructure"]),
    "衍生品": ("derivatives", ["derivatives"]),
    "风险管理": ("derivatives", ["derivatives", "institutional-risk"]),
    "交易后基础设施": ("financial-markets", ["financial-markets", "institutional-risk"]),
    "机构风险": ("institutional-risk", ["institutional-risk"]),
    "区块链": ("blockchain", ["blockchain"]),
    "TRON": ("blockchain", ["blockchain", "tron-ecosystem"]),
    "TRON生态": ("blockchain", ["blockchain", "tron-ecosystem"]),
    "Crypto市场结构": ("crypto-market-structure", ["crypto-market-structure"]),
    "Crypto合规": ("regulation-compliance", ["regulation-compliance", "crypto-market-structure"]),
    "预测市场": ("prediction-outcome-markets", ["prediction-outcome-markets"]),
    "预测市场监管": ("prediction-outcome-markets", ["prediction-outcome-markets", "regulation-compliance"]),
    "预测市场数据": ("prediction-outcome-markets", ["prediction-outcome-markets"]),
    "产业战略": ("crypto-market-structure", ["industry-strategy", "crypto-market-structure"]),
    "事件市场候选概念": ("prediction-outcome-markets", ["prediction-outcome-markets", "industry-strategy"]),
}

# Term-level folder overrides (folder only; domains stay from category unless given)
FOLDER_OVERRIDES = {
    "Exchange": ("exchanges", ["exchanges", "financial-markets"]),
    "Venue": ("exchanges", ["exchanges", "financial-markets"]),
    "Stablecoin": ("stablecoins-wallets-payments", ["stablecoins-wallets-payments", "blockchain"]),
    "Wallet": ("stablecoins-wallets-payments", ["stablecoins-wallets-payments", "blockchain"]),
}

# Canonical-merge aliases: task-list names that are the SAME concept as a workbook
# term. Kept here so re-runs stay idempotent and no duplicate page is created.
EXTRA_ALIASES = {
    "Prime Brokerage": ["Prime Broker", "主经纪商"],
    "OTC": ["Institutional OTC", "场外交易"],
    "Order Flow Network Effect": ["Liquidity Network Effects", "流动性网络效应"],
    "Market Maker Incentive": ["Market Maker Incentives", "做市商激励"],
    "Custody": ["Custodian", "托管人"],
    "Distribution": ["Wallet Distribution", "钱包分发"],
    "CLOB": ["Central Limit Order Book"],
    "VaR": ["Value at Risk"],
    "AMM": ["Automated Market Maker"],
    "RFQ": ["Request for Quote"],
    "KYT": ["Know Your Transaction"],
    "KYC": ["Know Your Customer"],
    "AML": ["Anti-Money Laundering"],
    "PoW": ["Proof of Work"],
    "PoS": ["Proof of Stake"],
    "DPoS": ["Delegated Proof of Stake"],
    "L1": ["Layer 1"],
    "L2": ["Layer 2"],
    "EVM": ["Ethereum Virtual Machine"],
    "CEX": ["Centralized Exchange"],
    "DEX": ["Decentralized Exchange"],
    "BTTC": ["BitTorrent Chain"],
    "Fully Collateralized": ["Fully Collateralized Market"],
}

# Tier-1 terms (deterministic list, curated once here)
TIER1_TERMS = {
    "Prediction Market", "Event Contract", "Outcome Market", "Outcome Token",
    "Resolution", "Resolution Source", "Oracle Risk", "Resolution Risk",
    "Contract Semantics", "Contract Equivalence", "Canonical Event ID",
    "Event Risk", "Event VaR", "Fully Collateralized", "Implied Probability",
    "Market Integrity", "Inside Information",
    "CLOB", "Order Book", "Liquidity", "Depth", "Slippage", "Spread",
    "Adverse Selection", "Market Maker", "Basis Risk", "Hedging",
    "Margin", "Cross Margin", "Settlement", "Clearing", "Counterparty Risk",
    "Oracle", "Stablecoin", "Custody", "Derivative", "Binary/Digital Option",
    "Financial Markets", "Exchange", "Price Discovery", "Distribution",
    "Settlement Rail", "Risk Engine", "Policy Engine", "Data Infrastructure",
    "Liquidity Risk", "Concentration Risk", "Regulatory Risk", "Jurisdiction",
}

URL_SOURCE_TYPE = {
    "www.cftc.gov": ("regulatory-announcement", "CFTC"),
    "docs.polymarket.com": ("official-documentation", "Polymarket"),
    "developers.tron.network": ("official-documentation", "TRON DAO"),
    "tron.network": ("official-documentation", "TRON DAO"),
    "justlend.org": ("official-documentation", "JustLend DAO"),
    "bt.io": ("official-documentation", "BitTorrent Chain"),
    "winklink.org": ("official-documentation", "WINkLink"),
    "hyperliquid.gitbook.io": ("official-documentation", "Hyperliquid"),
    "www.cmegroup.com": ("official-documentation", "CME Group"),
}


def read_workbook():
    import openpyxl
    wb = openpyxl.load_workbook(SEED_XLSX, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        rows = []
        for row in wb[name].iter_rows(values_only=True):
            if any(c is not None and str(c).strip() for c in row):
                rows.append([str(c).strip() if c is not None else "" for c in row])
        sheets[name] = rows
    return sheets


def term_slug(term: str) -> str:
    if term in ENTITY_TERMS:
        return ENTITY_TERMS[term][3]
    return SLUG_OVERRIDES.get(term, slugify(term))


def valid_http_url(url: str) -> bool:
    """Accept only absolute HTTP(S) URLs without whitespace."""
    if not url or any(ch.isspace() for ch in url):
        return False
    parsed = urlparse(url)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def url_source_note(url: str) -> tuple[str, Path, str]:
    """Return (source_id, path, markdown) for a workbook-referenced URL."""
    if not valid_http_url(url):
        raise ValueError(f"invalid absolute HTTP(S) URL: {url}")
    host = urlparse(url).netloc
    stype, publisher = URL_SOURCE_TYPE.get(host, ("official-documentation", host))
    tail = urlparse(url).path.rstrip("/").split("/")[-1] or host
    slug = slugify(f"{publisher}-{tail}")[:60]
    sid = f"source:2026-08-26-{slug}"
    fname = f"src-2026-08-26-{slug}.md"
    candidate = vault_root() / "07_RESEARCH" / "sources" / fname
    if candidate.exists():
        existing = load_note(candidate)
        if existing.frontmatter.get("url") != url:
            suffix = hashlib.sha256(url.encode("utf-8")).hexdigest()[:8]
            slug = f"{slug[:51]}-{suffix}"
            sid = f"source:2026-08-26-{slug}"
            fname = f"src-2026-08-26-{slug}.md"
    fm = {
        "id": sid,
        "type": "source",
        "source_type": stype,
        "title": f"{publisher}: {tail}",
        "publisher": publisher,
        "author": "",
        "published_at": None,
        "accessed_at": TODAY,
        "url": url,
        "primary_source": True,
        "reliability": "high",
        "content_hash": "",
        "archive_path": "",
        "status": "seed",
        "importance": "tier-3",
        "domains": [],
        "tags": ["source"],
        "created": TODAY,
        "updated": TODAY,
        "confidence": "medium",
        "epistemic_status": "confirmed",
        "confidentiality": "public-source",
        "sources": [],
        "related": [],
        "import_origin": MARKER,
    }
    body = f"""
# Source: {publisher} — {tail}

## What This Source Is | 来源是什么

来自学习地图 workbook 的引用 URL: <{url}>

该 URL 由 workbook 携带, 导入时**未独立抓取验证内容** — 引用它的断言 confidence 上限为 medium, 直到 researcher 实际访问并回填 `content_hash` 与摘录。

## Key Claims Extracted | 提取的关键断言

| 断言 | 档位 | 被哪些页引用 |
|---|---|---|
| (待 researcher 回填) | | |

## Reliability Notes | 可靠性备注

域名 `{host}` 属于 {publisher} — {'监管机构一手文件' if stype == 'regulatory-announcement' else '官方文档'}, 来源优先级高。

## Freshness | 时效

- accessed: {TODAY} (仅登记, 未抓取)
- last verified: —
- suggested review: 2026-11-26
"""
    return sid, vault_root() / "07_RESEARCH" / "sources" / fname, build_frontmatter(fm) + body


def concept_page(row: list[str]) -> tuple[Path, str, str, str]:
    """Return (path, markdown, slug, kind) for one 术语总表 row."""
    term, zh, category, explain, why, example, application, url = (row + [""] * 8)[:8]
    slug = term_slug(term)
    is_entity = term in ENTITY_TERMS
    if is_entity:
        prefix, folder, ntype, _ = ENTITY_TERMS[term]
        subfolder, domains = None, ["blockchain", "tron-ecosystem"]
        path = vault_root() / folder / f"{slug}.md"
    else:
        prefix, ntype = "concept", "concept"
        subfolder, domains = CATEGORY_MAP.get(category, ("financial-markets", ["financial-markets"]))
        if term in FOLDER_OVERRIDES:
            subfolder, domains = FOLDER_OVERRIDES[term]
        path = vault_root() / "02_CONCEPTS" / subfolder / f"{slug}.md"

    nid = f"{prefix}:{slug}"
    aliases = [a for a in {term, zh.split("/")[0].strip()} if a and a.lower() != slug.replace("-", " ")]
    aliases += [a for a in EXTRA_ALIASES.get(term, []) if a not in aliases]
    src_ids = [WORKBOOK_SOURCE_ID]
    if valid_http_url(url):
        sid, _, _ = url_source_note(url)
        src_ids.append(sid)

    fm = {
        "id": nid,
        "type": ntype,
        "title": term,
        "title_zh": zh,
        "title_en": term,
        "aliases": sorted(aliases),
        "status": "seed",
        "importance": "tier-1" if term in TIER1_TERMS else "tier-2",
        "domains": domains,
        "tags": [ntype, "xlsx-import"],
        "created": TODAY,
        "updated": TODAY,
        "last_verified": None,
        "review_after": "2027-02-26",
        "confidence": "medium",
        "epistemic_status": "mixed",
        "confidentiality": "internal",
        "sources": src_ids,
        "related": [],
        "prerequisites": [],
        "import_origin": MARKER,
        "import_category": category,
    }

    sec = [f"\n# {term} | {zh}\n"]
    sec.append("> 本页由学习地图 workbook 确定性导入 (status: seed)。原文字段完整保留; enrich 时不删原文, 追加即可。\n")
    if explain:
        sec.append(f"## Executive Definition / Chinese Explanation | 定义与解释\n\n{explain}\n")
    if why:
        sec.append(f"## Why This Matters | 为什么重要\n\n{why}\n")
    if example:
        sec.append(f"## Concrete Example | 例子\n\n{example}\n")
    if application:
        sec.append(f"## In Practice | 实战里怎么用\n\n{application}\n")
    sec.append("## Related Concepts | 相关概念\n\n- (librarian 待补链)\n")
    qs = [f"- Q: 不看笔记, 用两三句话向一个聪明的外行解释 {term}。\n  A: 见上文定义。"]
    if application:
        qs.append(f"- Q: {term} 在真实交易或尽调里怎么用?\n  A: 见上文 In Practice。")
    sec.append("## Active-Recall Questions | 主动回忆题\n\n" + "\n".join(qs) + "\n")
    src_lines = ["- [[src-2026-08-26-industry-learning-map-xlsx]] — 学习地图 workbook (user-direct)"]
    if valid_http_url(url):
        _, source_path, _ = url_source_note(url)
        src_lines.append(f"- [[{source_path.stem}]] — <{url}>")
    sec.append("## Sources\n\n" + "\n".join(src_lines) + "\n")
    sec.append(f"<!-- timeline -->\n\n## Timeline\n\n- **{TODAY}** — 从学习地图 workbook 导入 (分类: {category})。 [Source: [[src-2026-08-26-industry-learning-map-xlsx]]]\n")

    return path, build_frontmatter(fm) + "\n" + "\n".join(sec), slug, ntype


def verbatim_page(nid, path_rel, title, title_zh, header, rows, note, domains, folder_note=""):
    fm = {
        "id": nid, "type": "report", "title": title, "title_zh": title_zh,
        "aliases": [], "status": "seed", "importance": "tier-1",
        "domains": domains, "tags": ["xlsx-import"],
        "created": TODAY, "updated": TODAY, "confidence": "medium",
        "epistemic_status": "mixed", "confidentiality": "internal",
        "sources": [WORKBOOK_SOURCE_ID], "related": [], "import_origin": MARKER,
    }
    lines = [f"\n# {title_zh}\n", f"> {note}\n"]
    lines.append("| " + " | ".join(header) + " |")
    lines.append("|" + "---|" * len(header))
    for r in rows:
        cells = [(c or "—").replace("|", "\\|").replace("\n", " ") for c in (r + [""] * len(header))[:len(header)]]
        lines.append("| " + " | ".join(cells) + " |")
    lines.append("\n## Sources\n\n- [[src-2026-08-26-industry-learning-map-xlsx]]\n")
    lines.append(f"<!-- timeline -->\n\n## Timeline\n\n- **{TODAY}** — 从学习地图 workbook 原文导入。\n")
    return vault_root() / path_rel, build_frontmatter(fm) + "\n" + "\n".join(lines)


def workbook_source_note() -> tuple[Path, str]:
    h = sha256_file(SEED_XLSX)
    fm = {
        "id": WORKBOOK_SOURCE_ID, "type": "source", "source_type": "user-direct",
        "title": "行业术语与市场结构学习地图 (Excel workbook)",
        "publisher": "Irene Sun (with ChatGPT assistance)", "author": "Irene Sun",
        "published_at": None, "accessed_at": TODAY,
        "url": "", "primary_source": True, "reliability": "high",
        "content_hash": h, "archive_path": "_seed/learning-map.xlsx",
        "status": "reviewed", "importance": "tier-1", "domains": [],
        "tags": ["source"], "created": TODAY, "updated": TODAY,
        "confidence": "high", "epistemic_status": "mixed",
        "confidentiality": "internal", "sources": [], "related": [],
        "import_origin": MARKER,
    }
    body = f"""
# Source: 行业术语与市场结构学习地图

## What This Source Is | 来源是什么

Irene 提供的 4-sheet Excel workbook (术语总表 134 条 / 学习地图 8 阶段 / 生态游戏版图 13 角色 / 机构对话速查 5 对象)。这是 vault 的**初始 seed 语料**, 属 user-direct 来源 — 代表 Irene 已消化认可的知识框架。

- SHA-256: `{h}`
- 原件: 由 `$IBRAIN_WORKBOOK` 指定 (只读保存, 不入库)

## Reliability Notes | 可靠性备注

user-direct 优先级最高, 但其中的**行业事实**仍按 workbook 内嵌 URL / 后续 research 独立核验; workbook 里的战略判断按 hypothesis/analysis 处理, 不自动升级为 decision。

## Freshness

- accessed: {TODAY}
- suggested review: 2027-02-26 (预测市场监管在快速变化, CFTC 提案落地时提前复核)
"""
    return vault_root() / "07_RESEARCH" / "sources" / "src-2026-08-26-industry-learning-map-xlsx.md", build_frontmatter(fm) + body


def write_page(path: Path, content: str, report: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        existing = load_note(path)
        origin = existing.frontmatter.get("import_origin")
        if origin != MARKER:
            report["skipped_human_owned"].append(str(path.relative_to(vault_root())))
            return
        if path.read_text(encoding="utf-8") == content:
            report["unchanged"].append(str(path.relative_to(vault_root())))
            return
        report["updated"].append(str(path.relative_to(vault_root())))
    else:
        report["created"].append(str(path.relative_to(vault_root())))
    path.write_text(content, encoding="utf-8")


def main() -> int:
    if not SEED_XLSX.exists():
        print(f"ERROR: workbook not found at {SEED_XLSX}")
        return 1
    sheets = read_workbook()
    report = {"created": [], "updated": [], "unchanged": [], "skipped_human_owned": [],
              "missing_values": [], "invalid_urls": [], "entity_terms": [], "rows_total": 0}

    # workbook source note + URL source notes
    p, c = workbook_source_note()
    write_page(p, c, report)
    seen_urls = {}
    terms = sheets.get("术语总表", [])[1:]
    report["rows_total"] = len(terms)
    for row in terms:
        url = (row + [""] * 8)[7]
        if url:
            if not valid_http_url(url):
                report["invalid_urls"].append(f"{row[0]}: {url}")
                continue
            if url not in seen_urls:
                sid, sp, sc = url_source_note(url)
                seen_urls[url] = sid
                write_page(sp, sc, report)

    # concept pages
    slugs_seen = {}
    for row in terms:
        term = row[0]
        row8 = (row + [""] * 8)[:8]
        for idx, fname in [(1, "中文"), (3, "通俗解释")]:
            if not row8[idx]:
                report["missing_values"].append(f"{term}: 缺 {fname}")
        slug = term_slug(term)
        if slug in slugs_seen:
            report["missing_values"].append(f"DUPLICATE SLUG {slug}: {term} vs {slugs_seen[slug]}")
            continue
        slugs_seen[slug] = term
        path, content, slug, ntype = concept_page(row8)
        if ntype != "concept":
            report["entity_terms"].append(f"{term} → {path.relative_to(vault_root())}")
        write_page(path, content, report)

    # verbatim sheets
    lm = sheets.get("学习地图", [])
    if lm:
        p, c = verbatim_page("curr:curriculum-source-map", "10_LEARNING/curriculum/curriculum-source-map.md",
                             "Curriculum Source Map (workbook)", "学习地图 · workbook 原文",
                             lm[0], lm[1:],
                             "8 阶段学习地图原文。手写 curriculum 以此为骨架; 本页保持与 workbook 一致, 不手改。",
                             ["learning"])
        write_page(p, c, report)
    eco = sheets.get("生态游戏版图", [])
    if eco:
        p, c = verbatim_page("report:ecosystem-roles-map", "06_RELATIONSHIPS/ecosystem-maps/ecosystem-roles-map.md",
                             "Ecosystem Roles Map (workbook)", "生态游戏版图 · workbook 原文",
                             eco[0], eco[1:],
                             "13 个生态角色: 钱在哪、最怕什么、怎么结算、以及一个中立数据层应当如何与其相处。",
                             ["industry-strategy", "crypto-market-structure"])
        write_page(p, c, report)
    cheat = sheets.get("机构对话速查", [])
    if cheat:
        p, c = verbatim_page("report:institutional-conversation-cheatsheet",
                             "10_LEARNING/expert-questions/institutional-conversation-cheatsheet.md",
                             "Institutional Conversation Cheatsheet (workbook)", "机构对话速查 · workbook 原文",
                             cheat[0], cheat[1:],
                             "见机构人士之前 5 分钟过一遍: 对方的钱在哪里、最怕什么、该问什么。",
                             ["learning", "institutional-risk"])
        write_page(p, c, report)

    # report
    rep_lines = [
        "# XLSX Import Report — 学习地图 workbook",
        "",
        f"- 运行时间: {TODAY}",
        f"- 输入: `$IBRAIN_WORKBOOK` (SHA-256 见 import-manifest)",
        f"- 术语总表行数: {report['rows_total']}",
        f"- 新建页面: {len(report['created'])}",
        f"- 重写页面 (importer-owned): {len(report['updated'])}",
        f"- 未变: {len(report['unchanged'])}",
        f"- 跳过 (已被人工接管, 需人工 merge): {len(report['skipped_human_owned'])}",
        f"- 判定为实体的术语 (落 03_ENTITIES): {len(report['entity_terms'])}",
        f"- 缺值 / 需人工复核: {len(report['missing_values'])}",
        f"- 无效 URL: {len(report['invalid_urls'])}",
        "",
        "## 判定为实体的术语",
        *([f"- {x}" for x in report["entity_terms"]] or ["- (无)"]),
        "",
        "## 缺值明细 (不阻塞导入, 仅提示)",
        *([f"- {x}" for x in report["missing_values"]] or ["- (无)"]),
        "",
        "## 无效 URL",
        *([f"- {x}" for x in report["invalid_urls"]] or ["- (无)"]),
        "",
        "## 需人工 merge (页面已被人工接管, workbook 行未写入)",
        *([f"- {x}" for x in report["skipped_human_owned"]] or ["- (无)"]),
        "",
        "## 幂等规则",
        "页面带 `import_origin: xlsx-learning-map` = importer 拥有, 重跑会同步 workbook 变更;",
        "人工 enrich 后请改为 `import_origin: xlsx-learning-map+manual`, importer 即不再覆盖。",
    ]
    rp = vault_root() / "90_META" / "import-reports" / "xlsx-import-report.md"
    rp.parent.mkdir(parents=True, exist_ok=True)
    rp.write_text("\n".join(rep_lines) + "\n", encoding="utf-8")
    print(f"ingest_xlsx: rows={report['rows_total']} created={len(report['created'])} "
          f"updated={len(report['updated'])} unchanged={len(report['unchanged'])} "
          f"skipped={len(report['skipped_human_owned'])}")
    print(f"report → {rp}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
