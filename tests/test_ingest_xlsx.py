import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))
import ingest_xlsx as ix


@pytest.fixture
def synthetic_workbook(tmp_path, monkeypatch):
    """A minimal 4-sheet workbook, so the importer is exercised in CI without the
    private seed corpus. Shape mirrors the real one; content is invented."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "术语总表"
    ws.append(["术语", "中文", "分类", "解释", "为什么重要", "例子", "应用", "URL"])
    ws.append(["CLOB", "中央限价订单簿", "市场微观结构", "解释文本。", "重要因为。", "例子。", "应用。", ""])
    ws.append(["TestTerm", "测试词", "预测市场", "解释文本。", "重要因为。", "例子。", "应用。", ""])
    for name in ("学习地图", "生态游戏版图", "机构对话速查"):
        s = wb.create_sheet(name)
        s.append(["列一", "列二"])
        s.append(["值一", "值二"])
    path = tmp_path / "learning-map.xlsx"
    wb.save(path)
    monkeypatch.setattr(ix, "SEED_XLSX", path)
    return path


def test_slug_overrides_stable():
    assert ix.term_slug("CLOB") == "central-limit-order-book"
    assert ix.term_slug("Perpetual Futures / Perp") == "perpetual-futures"
    assert ix.term_slug("Bandwidth (TRON)") == "tron-bandwidth"
    assert ix.term_slug("Liquidity") == "liquidity"


def test_entity_terms_routed_to_entities():
    assert ix.term_slug("WINkLink") == "winklink"
    assert "WINkLink" in ix.ENTITY_TERMS


@pytest.mark.skipif(not ix.SEED_XLSX.exists(),
                    reason="private seed workbook not present (expected in a public checkout)")
def test_category_map_covers_workbook():
    import openpyxl
    wb = openpyxl.load_workbook(ix.SEED_XLSX, read_only=True, data_only=True)
    ws = wb["术语总表"]
    cats = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not row or not row[0]:
            continue
        if row[2]:
            cats.add(str(row[2]).strip())
    unmapped = cats - set(ix.CATEGORY_MAP)
    assert not unmapped, f"unmapped categories: {unmapped}"


def test_concept_page_shape():
    row = ["TestTerm", "测试词", "预测市场", "解释文本。", "重要因为。", "例子。", "实战应用说明。", ""]
    path, content, slug, ntype = ix.concept_page(row)
    assert slug == "testterm" and ntype == "concept"
    assert "02_CONCEPTS/prediction-outcome-markets" in str(path)
    assert 'id: "concept:testterm"' in content  # colon-bearing scalars are quoted
    assert "解释文本。" in content and "实战应用说明。" in content
    assert "<!-- timeline -->" in content
    assert "import_origin: xlsx-learning-map" in content


def test_invalid_url_is_rejected_and_not_linked(monkeypatch, tmp_path):
    monkeypatch.setattr(ix, "vault_root", lambda: tmp_path)
    assert not ix.valid_http_url("http-not-a-url")
    assert not ix.valid_http_url("javascript:alert(1)")
    row = ["TestTerm", "测试词", "预测市场", "解释", "重要", "例子", "意义", "http-not-a-url"]
    _, content, _, _ = ix.concept_page(row)
    assert "http-not-a-url" not in content
    assert ix.WORKBOOK_SOURCE_ID in content


def test_source_id_collision_gets_hash_suffix(monkeypatch, tmp_path):
    monkeypatch.setattr(ix, "vault_root", lambda: tmp_path)
    first = "https://example.com/a/overview"
    second = "https://example.com/b/overview"
    sid1, path1, content1 = ix.url_source_note(first)
    path1.parent.mkdir(parents=True, exist_ok=True)
    path1.write_text(content1, encoding="utf-8")
    sid2, path2, _ = ix.url_source_note(second)
    assert sid1 != sid2
    assert path1 != path2
    assert len(path2.stem.rsplit("-", 1)[-1]) == 8


def test_full_import_is_idempotent(monkeypatch, tmp_path, synthetic_workbook):
    monkeypatch.setattr(ix, "vault_root", lambda: tmp_path)
    assert ix.main() == 0
    report = tmp_path / "90_META" / "import-reports" / "xlsx-import-report.md"
    first = {p.relative_to(tmp_path): p.read_bytes() for p in tmp_path.rglob("*.md") if p != report}
    assert ix.main() == 0
    second = {p.relative_to(tmp_path): p.read_bytes() for p in tmp_path.rglob("*.md") if p != report}
    assert first == second


def test_richer_existing_note_is_preserved(monkeypatch, tmp_path, synthetic_workbook):
    monkeypatch.setattr(ix, "vault_root", lambda: tmp_path)
    path = tmp_path / "02_CONCEPTS" / "market-microstructure" / "central-limit-order-book.md"
    path.parent.mkdir(parents=True)
    original = "---\nid: concept:central-limit-order-book\ntype: concept\n---\nRicher human note\n"
    path.write_text(original, encoding="utf-8")
    assert ix.main() == 0
    assert path.read_text(encoding="utf-8") == original
